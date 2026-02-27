from cogsol.tools import BaseTool, tool_params


class QueryExcelAttachment(BaseTool):
    """Script tool that reads an Excel file attached to the chat."""

    name = "query_excel_attachment"
    description = (
        "Read an Excel file that the user attached to the conversation. "
        "Returns the spreadsheet data as a formatted table."
    )
    show_tool_message = False

    @tool_params(
        sheet={
            "description": (
                "Worksheet name to read (e.g. 'Sheet1'). "
                "If not provided, reads the first sheet."
            ),
            "type": "string",
            "required": False,
        },
        cell_range={
            "description": (
                "Cell range to read (e.g. 'A1:D10'). "
                "If not provided, reads all data in the sheet."
            ),
            "type": "string",
            "required": False,
        },
    )
    def run(self, sheet="", cell_range="", chat=None, data=None, secrets=None, log=None):
        import io
        from django.apps import apps

        MessageAttachment = apps.get_model("assistant", "MessageAttachment")

        try:
            last_msg = (
                chat.messages.filter(role="user")
                .order_by("msg_num")
                .last()
                .get_message_dict()
            )
            attachments = last_msg.get("attachments", [])
            if not attachments:
                return "No file attached. Please upload an Excel file and try again."

            attachment_ref = attachments[0]
            if len(attachment_ref.split("-")) > 1:
                attachment_id = attachment_ref.split("-")[1]
            else:
                attachment_id = attachment_ref

            attachment = MessageAttachment.objects.get(id=attachment_id)
            file_bytes = attachment.load_file()
            log.append(f"File loaded: {attachment.content_type}")

        except Exception as e:
            log.append(f"Attachment error: {str(e)}")
            return f"Could not read the attached file: {str(e)}"

        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

            if sheet:
                if sheet not in wb.sheetnames:
                    return (
                        f"Sheet '{sheet}' not found. "
                        f"Available sheets: {', '.join(wb.sheetnames)}"
                    )
                ws = wb[sheet]
            else:
                ws = wb.active

            if cell_range:
                rows = ws[cell_range]
            else:
                rows = ws.iter_rows()

            table_rows = []
            for row in rows:
                table_rows.append(
                    [str(cell.value) if cell.value is not None else "" for cell in row]
                )

            if not table_rows:
                return "The selected range is empty."

            header = table_rows[0]
            separator = ["-" * max(len(h), 3) for h in header]
            lines = ["| " + " | ".join(header) + " |"]
            lines.append("| " + " | ".join(separator) + " |")
            for row in table_rows[1:]:
                padded = row + [""] * (len(header) - len(row))
                lines.append("| " + " | ".join(padded[:len(header)]) + " |")

            log.append(f"Parsed {len(table_rows)} rows from sheet '{ws.title}'")
            return f"Data from sheet **{ws.title}**:\n\n" + "\n".join(lines)

        except Exception as e:
            log.append(f"Parse error: {str(e)}")
            return f"Could not parse the Excel file: {str(e)}"
